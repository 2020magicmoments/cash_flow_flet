import flet as ft
from datetime import datetime
from data_manager import get_filtered_report

# Libraries for exporting
from fpdf import FPDF
import openpyxl

def ReportView(page):
    # --- STATE VARIABLES ---
    today = datetime.now()
    first_day = today.replace(day=1)
    
    txt_start = ft.TextField(value=first_day.strftime("%Y-%m-%d"), read_only=True, width=100, text_style=ft.TextStyle(size=12))
    txt_end = ft.TextField(value=today.strftime("%Y-%m-%d"), read_only=True, width=100, text_style=ft.TextStyle(size=12))
    
    list_container = ft.Column(scroll=ft.ScrollMode.AUTO, expand=True)

    # --- EXPORT LOGIC ---
    def save_file_result(e: ft.FilePickerResultEvent):
        if not e.path:
            return # User cancelled

        file_path = e.path
        rows = get_filtered_report(txt_start.value, txt_end.value)

        try:
            # === EXCEL EXPORT ===
            if file_path.endswith(".xlsx"):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Report"
                # Headers
                ws.append(["ID", "Type", "Category", "Amount", "Date"])
                # Data
                for row in rows:
                    ws.append(list(row))
                
                wb.save(file_path)
                page.snack_bar = ft.SnackBar(ft.Text(f"Saved Excel to {file_path}"), bgcolor="green")

            # === PDF EXPORT ===
            elif file_path.endswith(".pdf"):
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", size=12)
                
                # Title
                pdf.set_font("Arial", 'B', 16)
                pdf.cell(200, 10, txt="Transaction Report", ln=1, align='C')
                
                # Date Range
                pdf.set_font("Arial", size=10)
                pdf.cell(200, 10, txt=f"From: {txt_start.value}  To: {txt_end.value}", ln=1, align='C')
                pdf.ln(10)

                # Headers
                pdf.set_font("Arial", 'B', 10)
                pdf.cell(30, 10, "Type", 1)
                pdf.cell(60, 10, "Category", 1)
                pdf.cell(40, 10, "Amount", 1)
                pdf.cell(50, 10, "Date", 1)
                pdf.ln()

                # Rows
                pdf.set_font("Arial", size=10)
                for row in rows:
                    # row: (id, type, category, amount, date)
                    t_type = row[1].title()
                    cat = row[2]
                    # Note: FPDF doesn't support Rupee symbol by default, using 'Rs.'
                    amt = f"Rs. {row[3]}" 
                    date = row[4][:10]

                    pdf.cell(30, 10, t_type, 1)
                    pdf.cell(60, 10, cat, 1)
                    pdf.cell(40, 10, amt, 1)
                    pdf.cell(50, 10, date, 1)
                    pdf.ln()

                pdf.output(file_path)
                page.snack_bar = ft.SnackBar(ft.Text(f"Saved PDF to {file_path}"), bgcolor="green")
            
            page.snack_bar.open = True
            page.update()

        except Exception as ex:
            page.snack_bar = ft.SnackBar(ft.Text(f"Error: {str(ex)}"), bgcolor="red")
            page.snack_bar.open = True
            page.update()

    # Create FilePicker
    file_picker = ft.FilePicker(on_result=save_file_result)
    # Add to overlay immediately
    if file_picker not in page.overlay:
        page.overlay.append(file_picker)

    # --- BUTTON HANDLERS ---
    def export_pdf(e):
        file_picker.save_file(
            dialog_title="Save Report PDF",
            file_name=f"report_{datetime.now().strftime('%Y%m%d')}.pdf",
            allowed_extensions=["pdf"]
        )

    def export_excel(e):
        file_picker.save_file(
            dialog_title="Save Report Excel",
            file_name=f"report_{datetime.now().strftime('%Y%m%d')}.xlsx",
            allowed_extensions=["xlsx"]
        )

    # --- DATE PICKER LOGIC (Existing) ---
    def handle_start_date(e):
        txt_start.value = e.control.value.strftime("%Y-%m-%d")
        txt_start.update()

    def handle_end_date(e):
        txt_end.value = e.control.value.strftime("%Y-%m-%d")
        txt_end.update()

    picker_start = ft.DatePicker(on_change=handle_start_date)
    picker_end = ft.DatePicker(on_change=handle_end_date)
    
    if picker_start not in page.overlay:
        page.overlay.append(picker_start)
        page.overlay.append(picker_end)

    # --- RENDER LOGIC (Existing) ---
    def load_data(e=None):
        rows = get_filtered_report(txt_start.value, txt_end.value)
        list_container.controls.clear()
        
        if not rows:
            list_container.controls.append(
                ft.Container(content=ft.Text("No records found.", color="grey"), alignment=ft.alignment.center, padding=20)
            )
        else:
            period_income = sum(r[3] for r in rows if r[1] == 'income')
            period_expense = sum(r[3] for r in rows if r[1] == 'expense')
            
            list_container.controls.append(
                ft.Container(
                    bgcolor="#222222", padding=15, border_radius=10,
                    content=ft.Row([
                        ft.Column([ft.Text("Income", size=10, color="grey"), ft.Text(f"₹{period_income:,.0f}", color="green")]),
                        ft.Column([ft.Text("Expense", size=10, color="grey"), ft.Text(f"₹{period_expense:,.0f}", color="red")]),
                        ft.Column([ft.Text("Saving", size=10, color="grey"), ft.Text(f"₹{period_income-period_expense:,.0f}", color="white")]),
                    ], alignment=ft.MainAxisAlignment.SPACE_AROUND)
                )
            )
            list_container.controls.append(ft.Container(height=10))

            current_month_key = None
            for row in rows:
                t_type, category, amount, date_str = row[1], row[2], row[3], row[4]
                try:
                    dt_obj = datetime.strptime(date_str[:10], "%Y-%m-%d")
                    month_key = dt_obj.strftime("%B %Y")
                except:
                    month_key = "Unknown"

                if month_key != current_month_key:
                    list_container.controls.append(
                        ft.Container(content=ft.Text(month_key, color="#AAAAAA", weight="bold", size=14), padding=ft.padding.only(top=15, bottom=5), bgcolor="black")
                    )
                    current_month_key = month_key

                is_income = (t_type == "income")
                item = ft.Container(
                    padding=12, bgcolor="#111111", border_radius=8, margin=ft.margin.only(bottom=5),
                    content=ft.Row([
                        ft.Row([
                            ft.Icon(ft.Icons.ARROW_DOWNWARD if is_income else ft.Icons.ARROW_UPWARD, color="green" if is_income else "red", size=16),
                            ft.Column([ft.Text(category, color="white", weight="bold"), ft.Text(date_str, color="grey", size=10)], spacing=2)
                        ]),
                        ft.Text(f"{'+' if is_income else '-'} ₹{amount:,.0f}", color="green" if is_income else "white", weight="bold")
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
                )
                list_container.controls.append(item)
        
        if list_container.page:
            list_container.update()

    # --- LAYOUT CONSTRUCTION ---
    
    # 1. Date Filters
    filter_row = ft.Row(
        controls=[
            ft.Column([ft.Text("From", size=10, color="grey"), ft.ElevatedButton(content=ft.Row([ft.Icon(ft.Icons.CALENDAR_TODAY, size=14), txt_start], spacing=5), style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=5), padding=10), bgcolor="#1A1A1A", color="white", on_click=lambda _: page.open(picker_start))]),
            ft.Column([ft.Text("To", size=10, color="grey"), ft.ElevatedButton(content=ft.Row([ft.Icon(ft.Icons.CALENDAR_TODAY, size=14), txt_end], spacing=5), style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=5), padding=10), bgcolor="#1A1A1A", color="white", on_click=lambda _: page.open(picker_end))]),
            ft.IconButton(icon=ft.Icons.SEARCH, bgcolor="white", icon_color="black", on_click=load_data)
        ],
        alignment=ft.MainAxisAlignment.SPACE_BETWEEN
    )

    # 2. Export Buttons
    export_row = ft.Row(
        [
            ft.Text("Export as:", color="grey", size=12),
            ft.OutlinedButton("PDF", icon=ft.Icons.PICTURE_AS_PDF, on_click=export_pdf, style=ft.ButtonStyle(color="white")),
            ft.OutlinedButton("Excel", icon=ft.Icons.TABLE_CHART, on_click=export_excel, style=ft.ButtonStyle(color="white")),
        ],
        alignment=ft.MainAxisAlignment.END
    )

    load_data()

    return ft.View(
        "/report",
        controls=[
            ft.Container(
                padding=20,
                content=ft.Column(
                    [
                        filter_row,
                        ft.Container(height=10),
                        export_row, # Added Export Buttons here
                        ft.Divider(color="#333333"),
                        list_container
                    ],
                    expand=True
                )
            )
        ],
        appbar=ft.AppBar(
            leading=ft.IconButton(ft.Icons.ARROW_BACK, on_click=lambda e: page.go("/")),
            title=ft.Text("Reports", weight="bold"),
            center_title=True,
            bgcolor="#111111",
        ),
        bgcolor="black"
    )